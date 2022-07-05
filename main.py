from functools import lru_cache
import pathlib
from typing import Type

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from report_сard_via_classes import Worker

INITIAL_ROW_OF_NAMES: int = 13
FINAL_ROW_OF_NAMES: int = 49
COLUMN_OF_NAMES: int = 2


@lru_cache
def get_workers(sheet: Type[Worksheet]) -> list:
    workers_list = list()

    # Iteration on column of names
    for col in sheet.iter_cols(min_row=INITIAL_ROW_OF_NAMES,
                               max_row=FINAL_ROW_OF_NAMES,
                               min_col=COLUMN_OF_NAMES,
                               max_col=COLUMN_OF_NAMES):
        for cell in col:
            if cell.value is not None:
                workers_list.append(Worker(cell.coordinate, sheet))
    return workers_list


def fill_all_workers(wokers_list: list):
    for worker in wokers_list:
        worker.fill_worker_line()


def save_file(file_name):
    report_card_file = pathlib.Path(file_name)
    backup_report_card_file = report_card_file.parent.joinpath(f'backup_{report_card_file.name}')
    wb = load_workbook(filename=report_card_file)
    wb.save(backup_report_card_file)
    for sheet in wb._sheets:
        fill_all_workers(get_workers(sheet))
    wb.save(report_card_file)

