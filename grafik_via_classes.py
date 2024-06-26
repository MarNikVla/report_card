import itertools
import pathlib
from collections import Counter
from copy import deepcopy
from functools import cached_property, partial

from openpyxl import load_workbook
from openpyxl.cell import Cell

DAYS_TO_REMOVE = (None, 'X', 'Х')


class Sheet:
    """
    represents sheet of excel file, methods and attributes of them
    """

    def __init__(self, sheet):
        self.sheet = sheet

    @staticmethod
    def _type_of_day(cell: Cell) -> str:
        """
        Defines type of day of the week by the specified color

        :param cell: openpyxl.cell.cell.Cell
        :return: str in ['РД', 'В', 'КД', 'П'], type of day
        """
        working_day = 'РД'  # Рабочий день
        weekend = 'В'  # Выходной
        short_day = 'КД'  # Короткий день
        holiday = 'П'  # Праздник

        color_standard_red = 'FFFF0000'
        color_standard_light_green = 'FF92D050'
        color_standard_orange = 'FFFFC000'

        if cell.fill.start_color.index == color_standard_red:
            return holiday
        if cell.fill.start_color.index == color_standard_light_green:
            return weekend
        if cell.fill.start_color.index == color_standard_orange:
            return short_day
        return working_day

    @cached_property
    def _work_days_matrix(self) -> list:
        """
        Create list of type of days in month (working_day, weekend, holiday, short_day)
        :return: list of str in ['РД', 'В', 'КД', 'П']
        example: ['П', 'В', 'В', 'РД', 'РД', 'РД', 'В', 'В', 'П', 'В', 'РД', 'РД', 'РД']
                meaning: first day of the month - 'П', second day of the month - 'В', etc
        """
        first_day_of_month = self.sheet['D12']
        last_day_of_month = self.sheet['AH31']
        work_days_matrix = list()
        for row in self.sheet.iter_rows(min_row=first_day_of_month.row,
                                        max_row=last_day_of_month.row,
                                        min_col=first_day_of_month.column,
                                        max_col=last_day_of_month.column):
            for cell in row:
                if cell.value not in DAYS_TO_REMOVE:
                    work_days_matrix.append(self._type_of_day(cell))
        return work_days_matrix


class Worker(Sheet):
    """
    represents each worker on the sheet, methods and attributes of them
    """

    def __init__(self, cell_index, sheet):
        super().__init__(sheet)
        self.cell = sheet[cell_index]
        self.name = self.cell.value

    # def name(self):
    #     return self.cell.value

    def __str__(self):
        return f'{self.cell.value}'

    @property
    def cells_range(self) -> list:
        """
        represents how many hours the worker worked, rested etc
        :return: list of cell range of worker filled in by user of program
        example: ['В', 'В', '8/20', '8/20', 'В', '20/', '/8   20/', '/8', 'В', 'В', '8/20' ...]
                meaning: first day of the month - 'В'(weekend), second day of the month - 'В', etc

        """
        cells_range = list()
        for row in self.sheet.iter_rows(min_row=self.cell.row,
                                        max_row=self.cell.row,
                                        min_col=self.cell.column + 2,
                                        max_col=self.cell.column + 32):
            for cell in row:
                if cell.value not in DAYS_TO_REMOVE:
                    cells_range.append(str(cell.value))
        return cells_range

    # @cached_property
    # def is_28_hours_week(self):
    #     """
    #     If worker have 28 hours week duration return - True
    #     Specified by yellow color of worker Cell
    #     :return: True|False
    #     """
    #     color_standard_yellow = 'FFFFFF00'
    #     return self.cell.fill.start_color.index == color_standard_yellow

    @property
    def _work_days_matrix(self):
        """
        Create list of type of days in month (working_day, weekend, holiday, short_day) for worker
        :return: list of str in ['РД', 'В', 'КД', 'П']
        example: ['П', 'В', 'В', 'РД', 'РД', 'РД', 'В', 'В', 'П', 'В', 'РД', 'РД', 'РД']
               meaning: first day of the month - 'П', second day of the month - 'В', etc
        """
        return super(Worker, self)._work_days_matrix[:len(self.cells_range)]

    @cached_property
    def _normalize_workdays(self) -> list:
        """
        To account for the days that are included in the norm of hours
        :return: list of days without days_to_remove
        """
        days_to_remove = ['ОТ', 'У', 'ДО', 'Б', 'К', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ', 'НОД', 'ОВ',
                          'ПР']
        normalize_workdays = deepcopy(self._work_days_matrix)
        for index, cell in enumerate(self.cells_range):
            if cell in days_to_remove:
                normalize_workdays[index] = 0
        return normalize_workdays

    @cached_property
    def norm_of_hours(self) -> float:
        """
        To account for the norm of hours for worker
        :return: quantity of hours
        """
        counter = Counter(self._normalize_workdays)
        duration_of_day = 8
        # if self.is_28_hours_week:
        #     duration_of_day = 5.6
        duration_of_short_day = duration_of_day - 1
        norm_of_hours = counter['РД'] * duration_of_day + counter['КД'] * duration_of_short_day
        return round(norm_of_hours, 1)

    @cached_property
    def counter_of_days(self):
        '''
        get all days of worker in month
        :return: Counter
        example: Counter({'В': 12, '8/20': 8, '/8': 4, '20/': 3, '/8 20/': 3, '0/8 20/': 1})
        '''
        return Counter(self.cells_range)

    def get_weekends(self):
        return self.counter_of_days['В']

    def get_vacation_days(self):
        return self.counter_of_days['ОТ']

    def get_medical_days(self):
        return self.counter_of_days['Б']

    def get_other_days_off(self):
        return sum(
            [self.counter_of_days[key] for key in
             ['ОВ', 'У', 'ДО', 'К', 'ПР', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ', 'НОД']])

    def get_attendance_days(self):
        attendance_days = sum(self.counter_of_days.values()) - \
                          (self.get_weekends() + self.get_vacation_days() +
                           self.get_medical_days() + self.get_other_days_off())
        return attendance_days

    @staticmethod
    def floated_cells(cell_list: list[str]) -> list[str, float]:
        """
        try to float cell data in cell list
        :param cell_list: splited_cell_list
        :return: floated_cells_list
        """
        for i, cell in enumerate(cell_list):
            try:
                cell_list[i] = float(cell.replace(",", "."))
            except ValueError:
                pass
        return cell_list

    def split_cells(self, cells: list[str]) -> list[str, float]:
        """
        Prepared cells for further hours calculation
        :param cells: raw list[str]
        :return: prepared list[str]
        """
        splited_cell_list = []
        for cell in cells:
            if len(cell) > 1:
                splited_cell_list.extend(cell.split())
            else:
                splited_cell_list.append(cell)

        splited_cell_list = self.floated_cells(splited_cell_list)
        return splited_cell_list

    def count_hours(self, cells_range: list[str]) -> dict:
        """
        count hours from cells_range
        :param cells_range: raw list of cells.value
        :return: dict
        """
        all_hours = 0
        night_hours = 0
        splited_cell = self.split_cells(cells_range)

        counter_of_hours = Counter(splited_cell)
        for i in counter_of_hours.keys():
            if isinstance(i, float):
                all_hours += i * counter_of_hours[i]
            elif i == '8/20':
                all_hours += 12 * counter_of_hours[i]
            elif i in ['20/', '20/24']:
                all_hours += 4 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i in ['0/8', '/8']:
                all_hours += 8 * counter_of_hours[i]
                night_hours += 6 * counter_of_hours[i]
            elif i in ['8']:
                all_hours += 8 * counter_of_hours[i]
            elif i in ['4']:
                all_hours += 4 * counter_of_hours[i]
        return {'night_hours': round(night_hours, 1), 'all_hours': round(all_hours, 1)}

    def get_day_hours(self):
        count_day_hours = self.count_hours(self.cells_range)['all_hours']
        return count_day_hours

    def get_night_hours(self):
        count_night_hours = self.count_hours(self.cells_range)['night_hours']
        return count_night_hours

    def get_holidays_hours(self):
        holidays_range = list()
        for i, cell in enumerate(self.cells_range):
            if self._normalize_workdays[i] == 'П':
                holidays_range.append(cell)

        count_holiday_hours = self.count_hours(holidays_range)['all_hours']
        return count_holiday_hours

    def get_overwork(self):
        return round(self.get_day_hours() - self.norm_of_hours, 1)

    def fill_worker_line(self):
        """
        fill cells with calculation results
        :return:
        """
        # print(self)
        offset_row = 0
        # offset_column_attendance = 17
        offset_column_day_hours = 33
        offset_column_norm_hours = 34
        offset_column_night_hours = 35
        offset_column_holidays_hours = 36

        offset_column_overwork = 37

        cell_offset = partial(self.cell.offset, row=offset_row)
        cell_offset(column=offset_column_day_hours).value = self.get_day_hours() or None
        cell_offset(column=offset_column_norm_hours).value = self.norm_of_hours or None
        cell_offset(column=offset_column_night_hours).value = self.get_night_hours() or None
        cell_offset(column=offset_column_holidays_hours).value = self.get_holidays_hours() or None
        cell_offset(column=offset_column_overwork).value = self.get_overwork() or None


# for debug

# def save_filled_sheet(self):
#     self.fill_worker_line()
#     wb.save(BACKUP_REPORT_CARD_FILE)



# file_name2 = '(ГРАФИК РАБОТЫ) май.xlsx'
# report_card_file = pathlib.Path(file_name2)
# wb = load_workbook(filename=report_card_file)
# worker = Worker('B14', wb[wb.sheetnames[0]])
#
#
#
#
# print(worker.cell)
# print(worker.name)
#
# print(f'явки (дней): {worker.get_attendance_days()}')
# print(f'урочно (часов):{worker.get_day_hours()}')
# print(f'ночные (чачов):{worker.get_night_hours()}')
# print(f'праздничные (часов):{worker.get_holidays_hours()}')
# print(f'выходные (дней):{worker.get_weekends()}')
# print(f'отпуск (дней):{worker.get_vacation_days()}')
# print(f'болничный (дней):{worker.get_medical_days()}')
# print(f'прочие неявки (дней):{worker.get_other_days_off()}')
# print(f'переработка (часов):{worker.get_overwork()}')
# print(f'норма часов  {worker.norm_of_hours}')
# print(worker._work_days_matrix)
# print(worker._normalize_workdays)
