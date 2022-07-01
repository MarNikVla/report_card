import re
from collections import Counter
from copy import deepcopy
from functools import cached_property, partial

from openpyxl import load_workbook

REPORT_CARD_FILE = 'табель июнь ГТЦ.xlsx'
BACKUP_REPORT_CARD_FILE = f'backup_{REPORT_CARD_FILE}'

wb = load_workbook(filename=REPORT_CARD_FILE)
machinist_sheet_name, DEM_sheet_name, reason_sheet_name = wb.sheetnames

machinist_sheet, DEM_sheet, reason_of_absence_sheet = wb[machinist_sheet_name], \
                                                      wb[DEM_sheet_name], \
                                                      wb[reason_sheet_name]


class Sheet:
    def __init__(self, sheet):
        self.sheet = sheet

    @staticmethod
    def _type_of_day(cell):

        working_day = 'РД'  # Рабочий день
        weekend = 'В'  # Выходной
        short_day = 'КД'  # Короткий день
        holiday = 'П'  # Праздник

        color_standard_red = 'FFFF0000'
        color_standard_light_green = 'FF92D050'
        color_standard_orange = 'FFFFC000'

        if cell.fill.start_color.index == color_standard_red:
            return holiday
        if cell.fill.start_color.index == color_standard_light_green:
            return weekend
        if cell.fill.start_color.index == color_standard_orange:
            return short_day
        return working_day

    @cached_property
    def _work_days_matrix(self):
        first_day_of_month = self.sheet['C10']
        last_day_of_month = self.sheet['R11']
        work_days_matrix = list()
        for row in self.sheet.iter_rows(min_row=first_day_of_month.row,
                                        max_row=last_day_of_month.row,
                                        min_col=first_day_of_month.column,
                                        max_col=last_day_of_month.column):
            for cell in row:
                if cell.value not in (None, 'X'):
                    work_days_matrix.append(self._type_of_day(cell))
        # print('fsdfs', len(work_days_matrix), work_days_matrix)
        return work_days_matrix

    def __str__(self):
        return f'{self.sheet}'


class Worker(Sheet):
    def __init__(self, cell_index, sheet):
        super().__init__(sheet)
        self.cell = sheet[cell_index]

    def __str__(self):
        return f'{self.cell.value},row  {self.cell.row}, col {self.cell.column}'

    def __len__(self):
        return len(self.cells_range)

    @property
    def cells_range(self):
        cells_range = list()
        for col in self.sheet.iter_rows(min_row=self.cell.row,
                                        max_row=self.cell.row + 1,
                                        min_col=self.cell.column + 1,
                                        max_col=self.cell.column + 16):
            for cell in col:
                if cell.value not in (None, 'Х'):
                    cells_range.append(cell.value)
        return cells_range

    @cached_property
    def is_28_hours_week(self):
        # print(self.cell.fill.start_color.index)
        color_standard_yellow = 'FFFFFF00'
        return self.cell.fill.start_color.index == color_standard_yellow

    @property
    def _work_days_matrix(self):
        return super(Worker, self)._work_days_matrix[:len(self.cells_range)]

    @cached_property
    def _normalize_workdays(self):
        days_to_remove = ['ОТ', 'У', 'ДО', 'Б', 'К', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ', 'НОД']
        normalize_workdays = deepcopy(self._work_days_matrix)
        for index, cell in enumerate(self.cells_range):
            if cell in days_to_remove:
                normalize_workdays[index] = 0
        return normalize_workdays

    @cached_property
    def norm_of_hours(self):
        counter = Counter(self._normalize_workdays)
        duration_of_day = 8
        if self.is_28_hours_week:
            duration_of_day = 5.6
        duration_of_short_day = duration_of_day - 1
        norm_of_hours = counter['РД'] * duration_of_day + counter['КД'] * duration_of_short_day
        return round(norm_of_hours, 1)

    @cached_property
    def counter_of_days(self):
        return Counter(self._prepared_range)

    def get_weekends(self):
        return self.counter_of_days['В']

    def get_vacation_days(self):
        return self.counter_of_days['ОТ']

    def get_medical_days(self):
        return self.counter_of_days['Б']

    def get_other_days_off(self):
        return sum(
            [self.counter_of_days[key] for key in
             ['ОВ', 'У', 'ДО', 'К', 'ПР', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ', 'НОД']])

    def get_attendance_days(self):
        attendance_days = sum(self.counter_of_days.values()) - \
                          (self.get_weekends() + self.get_vacation_days() +
                           self.get_medical_days() + self.get_other_days_off())
        return attendance_days

    @property
    def _prepared_range(self):
        new_cells_list = list()
        for cell in self.cells_range:
            if cell is not None:
                try:
                    new_cell = float(cell.replace(",", "."))
                except ValueError:
                    # remove whitespaces
                    new_cell = re.sub(r'\s+', '', cell)
                new_cells_list.append(new_cell)
        # print(new_cells_list)
        return new_cells_list

    @staticmethod
    def count_hours(cells_range):
        day_hours = 0
        night_hours = 0
        counter_of_hours = Counter(cells_range)
        # print(counter_of_hours)

        for i in counter_of_hours.keys():
            if isinstance(i, float):
                day_hours += i * counter_of_hours[i]
            elif i == '8/20':
                day_hours += 12 * counter_of_hours[i]
            elif i in ['20/', '20/24']:
                day_hours += 4 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i in ['/820/24', '0/820/', '/820/']:
                day_hours += 12 * counter_of_hours[i]
                night_hours += 8 * counter_of_hours[i]
            elif i in ['820/24', '820/']:
                day_hours += 12 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i in ['420/24', '420/']:
                day_hours += 8 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i in ['0/8', '/8']:
                day_hours += 8 * counter_of_hours[i]
                night_hours += 6 * counter_of_hours[i]
        # d = {'night_hours': night_hours, 'all_hours': all_hours}
        return {'night_hours': round(night_hours, 1), 'day_hours': round(day_hours, 1)}

    def get_day_hours(self):
        count_day_hours = self.count_hours(self._prepared_range)['day_hours']
        return count_day_hours

    def get_night_hours(self):
        count_night_hours = self.count_hours(self._prepared_range)['night_hours']
        return count_night_hours

    def get_holidays_hours(self):
        holidays_range = list()
        for i, cell in enumerate(self._prepared_range):
            if self._normalize_workdays[i] == 'П':
                holidays_range.append(cell)
        count_holiday_hours = self.count_hours(holidays_range)['day_hours']
        return count_holiday_hours

    def get_overwork(self):
        # print('day_hours', self.get_day_hours())
        # print('norm', self.norm_of_hours)
        # print('holiday_hours', self.get_holidays_hours())
        return round(self.get_day_hours() - self.norm_of_hours - self.get_holidays_hours(), 1)

    def fill_worker_line(self):
        offset_row = 0
        offset_column_attendance = 17
        offset_column_day_hours = 18
        offset_column_night_hours = 20
        offset_column_holidays_hours = 21
        offset_column_weekends = 22
        offset_column_vacation = 23
        offset_column_medical = 24
        offset_column_other_days_off = 25
        offset_column_overwork = 26

        cell_offset = partial(self.cell.offset, row=offset_row)
        cell_offset(column=offset_column_attendance).value = self.get_attendance_days() or None
        cell_offset(column=offset_column_weekends).value = self.get_weekends() or None
        cell_offset(column=offset_column_vacation).value = self.get_vacation_days() or None
        cell_offset(column=offset_column_medical).value = self.get_medical_days() or None
        cell_offset(column=offset_column_other_days_off).value = self.get_other_days_off() or None
        cell_offset(column=offset_column_day_hours).value = self.get_day_hours() or None
        cell_offset(column=offset_column_night_hours).value = self.get_night_hours() or None
        cell_offset(column=offset_column_holidays_hours).value = self.get_holidays_hours() or None
        cell_offset(column=offset_column_overwork).value = self.get_overwork() or None
        # wb.save(BACKUP_REPORT_CARD_FILE)

    def save_filled_sheet(self):
        self.fill_worker_line()
        wb.save(BACKUP_REPORT_CARD_FILE)


# worker = Worker('B25', DEM_sheet)
# worker_women = Worker('B35', DEM_sheet)
# worker_women2 = Worker('B37', DEM_sheet)


# print(worker.get_medical_days())
# print(worker.get_other_days_off())
# print(worker.get_weekends())
# print(worker.get_overwork())
# print(worker.get_day_hours())
# print(worker.get_night_hours())
# print(worker.norm_of_hours)
# print(worker._normalize_workdays)
# print(worker._work_days_matrix)
# print(worker.save_filled_sheet())
