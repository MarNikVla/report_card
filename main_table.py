from functools import lru_cache
import pathlib
from typing import Type

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from report_сard_via_classes_new import Worker

INITIAL_ROW_OF_NAMES: int = 16
FINAL_ROW_OF_NAMES: int = 49
COLUMN_OF_NAMES: int = 2
MAKE_BACKUP = True


def get_initial_row(sheet: Type[Worksheet]) -> int:
    for row in sheet.iter_rows(min_col=COLUMN_OF_NAMES - 1,
                               max_col=COLUMN_OF_NAMES - 1,
                               values_only=True):
        pass
        # print(row)


@lru_cache
def get_workers(sheet: Type[Worksheet]) -> list[Worker]:
    """
    get all worker from sheet
    :param sheet: Excel sheet
    :return: list of Worker instance
    """
    workers_list = list()

    # Iteration on column of names
    for col in sheet.iter_cols(min_row=INITIAL_ROW_OF_NAMES,
                               max_row=FINAL_ROW_OF_NAMES,
                               min_col=COLUMN_OF_NAMES,
                               max_col=COLUMN_OF_NAMES, ):
        for cell in col:
            if cell.value is not None:
                workers_list.append(Worker(cell.coordinate, sheet))
    return workers_list


def fill_all_workers(workers_list: list[Worker]):
    """
    fill cells for all worker on sheet
    :param workers_list:
    :return: list[Worker]
    """
    for worker in workers_list:
        worker.fill_worker_line()


def make_backup(file_name):
    pass


def save_file(file_name):
    """
    save Excel file with calculation results and do backup
    :param file_name: name of exel file
    :return:
    """
    report_card_file = pathlib.Path(file_name)
    backup_report_card_file = report_card_file.parent.joinpath(f'backup_{report_card_file.name}')
    wb = load_workbook(filename=report_card_file)
    if MAKE_BACKUP:
        wb.save(backup_report_card_file)
    for sheet in wb._sheets:
        fill_all_workers(get_workers(sheet))
    wb.save(report_card_file)


# for debug

# file_name = 'табель январь ГТЦ.xlsx'
# save_file(file_name)

file_name = 'табель март ГТЦ 2.xlsx'
report_card_file = pathlib.Path(file_name)
wb = load_workbook(filename=report_card_file)
# initial_row = get_initial_row(wb[wb.sheetnames[0]])
# print(f'initial row: {initial_row}')


worker = Worker('B16', wb[wb.sheetnames[1]])
print(worker.name)
print(worker.get_day_hours())

print(worker.counter_of_days)
print(f'явки (дней): {worker.get_attendance_days()}')
print(f'урочно (часов):{worker.get_day_hours()}')
print(f'ночные (чачов):{worker.get_night_hours()}')
print(f'праздничные (часов):{worker.get_holidays_hours()}')
print(f'выходные (дней):{worker.get_weekends()}')
print(f'отпуск (дней):{worker.get_vacation_days()}')
print(f'болничный (дней):{worker.get_medical_days()}')
print(f'прочие неявки (дней):{worker.get_other_days_off()}')
print(f'переработка (часов):{worker.get_overwork()}')
print(f'переработка реальная (часов):{worker.get_overwork_real()}')
print(worker._work_days_matrix)
print(f'праздничные по приказу (часов):{worker.get_cells_range_work_by_order()}')

print(f'норма часов  {worker.norm_of_hours}')
print(worker._normalize_workdays)

# save_file(file_name)
