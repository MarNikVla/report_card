from functools import lru_cache
from itertools import compress
from typing import Type

from openpyxl import load_workbook
from collections import Counter
import re

from openpyxl.worksheet.worksheet import Worksheet

REPORT_CARD_FILE = 'test.xlsx'
BACKUP_REPORT_CARD_FILE = f'backup_{REPORT_CARD_FILE}'

wb = load_workbook(filename=REPORT_CARD_FILE)
machinist_sheet_name, DEM_sheet_name, reason_sheet_name = wb.sheetnames

machinist_sheet, DEM_sheet, reason_of_absence_sheet = wb[machinist_sheet_name], \
                                                      wb[DEM_sheet_name], \
                                                      wb[reason_sheet_name]


@lru_cache
def get_lines_of_working_days_dict(sheet: Type[Worksheet]) -> dict:
    initial_row_of_names = 13
    final_row_of_names = 49
    column_of_names = 2
    lines_of_working_days_dict = dict()

    # Iteration on column of names
    for col in sheet.iter_cols(min_row=initial_row_of_names,
                               max_row=final_row_of_names,
                               min_col=column_of_names,
                               max_col=column_of_names):
        for cell in col:
            if cell.value is not None:
                lines_of_working_days_dict[cell.coordinate] = \
                    [cell.offset(row=i, column=j).value for i in [0, 1] for j in range(1, 17)]
    # for debug
    # print(f'lines_of_working_days_dict: {lines_of_working_days_dict}')
    return lines_of_working_days_dict


def get_normalize_cells_list(raw_cells_list: list[str]) -> list:
    """

    :param raw_cells_list: list[str] of working days of one worker
    :return: list[str|float] str: without whitespaces or float type if it's possible and remove
    elem with 'Х' value
    """
    new_cells_list = list()
    for cell in raw_cells_list:
        if cell is not None:
            try:
                new_cell = float(cell.replace(",", "."))
            except ValueError:
                # remove whitespaces
                new_cell = re.sub(r'\s+', '', cell)
            new_cells_list.append(new_cell)
    # remove elem with 'Х'
    new_cells_list.remove('Х')
    return new_cells_list


def count_days(normalized_list: list) -> dict:
    """

    :param normalized_list: list returned from get_normalize_cells_list()
    :return:
    """
    counter = Counter(normalized_list)
    days_dict = dict()

    absence_days = counter['В']
    absence_paid_days = counter['НОД']
    vacation_days = counter['ОТ']
    medical_days = counter['Б']
    other_absence_days = sum(
        [counter[key] for key in ['ОВ', 'У', 'ДО', 'К', 'ПР', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ']])
    attendance_days = sum(
        counter.values()) - (absence_days + vacation_days + medical_days +
                             other_absence_days + absence_paid_days)

    days_dict['attendance_days'] = attendance_days or None
    days_dict['absence_days'] = absence_days or None
    days_dict['vacation_days'] = vacation_days or None
    days_dict['medical_days'] = medical_days or None
    days_dict['other_absence_days'] = other_absence_days or None
    days_dict['absence_paid_days'] = absence_paid_days or None

    return days_dict


def count_hours(normalized_list: list) -> dict:
    """

    :param normalized_list: list returned from get_normalize_cells_list()
    :return: dict of night hours, days hours of work
    """
    hours = 0
    night_hours = 0
    counter = Counter(normalized_list)
    hours_dict = dict()
    for i in counter.keys():
        if isinstance(i, float):
            hours += i * counter[i]
        elif i == '8/20':
            hours += 12 * counter[i]
        elif i in ['20/', '20/24']:
            hours += 4 * counter[i]
            night_hours += 2 * counter[i]
        elif i in ['/820/24', '0/820/', '/820/']:
            hours += 12 * counter[i]
            night_hours += 8 * counter[i]
        elif i == '820/':
            hours += 12 * counter[i]
            night_hours += 2 * counter[i]
        elif i == '420/':
            hours += 8 * counter[i]
            night_hours += 2 * counter[i]
        elif i in ['0/8', '/8']:
            hours += 8 * counter[i]
            night_hours += 6 * counter[i]

    hours_dict['hours'] = hours or None
    hours_dict['night_hours'] = night_hours or None

    return hours_dict


def fill_days_cells(sheet: Type[Worksheet], worker_cell_index: str):
    offset_row = 0
    offset_column_attendance = 17
    offset_column_absence = 22
    offset_column_vacation = 23
    offset_column_medical = 24
    offset_column_other_absence = 25
    offset_column_other_absence_paid = 26

    lines_of_working_days = get_lines_of_working_days_dict(sheet)[worker_cell_index]
    normalize_cells_list = get_normalize_cells_list(lines_of_working_days)
    days = count_days(normalize_cells_list)

    # filling cells
    # cell of attendance_days = cell_index.offset(row=0, column=17)
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_attendance).value = days[
        'attendance_days']
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_absence).value = days[
        'absence_days']
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_vacation).value = days[
        'vacation_days']
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_medical).value = days[
        'medical_days']
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_other_absence).value = \
        days[
            'other_absence_days']
    sheet[worker_cell_index].offset(row=offset_row,
                                    column=offset_column_other_absence_paid).value = \
        days[
            'absence_paid_days']


def fill_hours_cells(sheet: Type[Worksheet], worker_cell_index: str):
    offset_row = 0
    offset_column_hours = 18
    offset_column_night_hours = 20

    lines_of_worker_days_list = get_lines_of_working_days_dict(sheet)[worker_cell_index]
    normalize_cells_list = get_normalize_cells_list(lines_of_worker_days_list)
    hours = count_hours(normalize_cells_list)

    # filling cells
    # cell of hours = cell_index.offset(row=0, column=18)
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_hours).value = hours[
        'hours']
    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_night_hours).value = hours[
        'night_hours']


def fill_holidays_cell(sheet, worker_cell_index: str):
    offset_row = 0
    offset_column_holidays_hours = 21

    matrix_of_holidays = get_matrix_of_holidays(sheet)  # example: [0,0,0,0,1,0,0,...,0]
    normalize_cells_list = get_normalize_cells_list(
        get_lines_of_working_days_dict(sheet)[worker_cell_index])

    hours_on_holidays_dict = count_hours(list(compress(normalize_cells_list, matrix_of_holidays)))

    # filling cells
    # cell of hours = cell_index.offset(row=0, column=18)
    sheet[worker_cell_index].offset(row=offset_row,
                                    column=offset_column_holidays_hours).value = \
        hours_on_holidays_dict['hours']


def fill_overworking_cell(sheet, worker_cell_index: str):
    """
    TODO
    not working correct
    :param sheet:
    :param worker_cell_index:
    :return:
    """
    offset_row = 0
    offset_column_overworking = 27

    lines_of_worker_days_list = get_lines_of_working_days_dict(sheet)[worker_cell_index]
    normalize_cells_list = get_normalize_cells_list(lines_of_worker_days_list)
    hours = count_hours(normalize_cells_list)

    absence = count_days(normalize_cells_list)
    # filling cells
    # cell of hours = cell_index.offset(row=0, column=18)

    sheet[worker_cell_index].offset(row=offset_row, column=offset_column_overworking).value = \
        hours['hours'] - get_norm_of_hours_dict(sheet)['AH8'] - get_hours_on_holidays(sheet,
                                                                                      worker_cell_index)

    # wb.save(BACKUP_REPORT_CARD_FILE)


def save_sheet(sheet: Type[Worksheet]):
    wb.save(BACKUP_REPORT_CARD_FILE)
    lines_of_working_days_dict = get_lines_of_working_days_dict(sheet)
    for worker_cell_index in lines_of_working_days_dict.keys():
        fill_days_cells(sheet, worker_cell_index)
        fill_hours_cells(sheet, worker_cell_index)
        fill_holidays_cell(sheet, worker_cell_index)
        # fill_overworking_cell(sheet, worker_cell_index)

    # write_days_to_file(sheet):
    wb.save(REPORT_CARD_FILE)


def get_matrix_of_holidays(sheet: Type[Worksheet]) -> list:
    color_standard_red = 'FFFF0000'
    matrix_of_holidays = list()
    min_row = 10
    min_col = 3
    for col in sheet.iter_rows(min_row=min_row,
                               max_row=min_row + 1,
                               min_col=min_col,
                               max_col=min_col + 15):
        for cell in col:

            if cell.value == 'Х':
                pass
            # if cell.color == standard.red (holiday)
            elif cell.fill.start_color.index == color_standard_red:
                matrix_of_holidays.append(1)
            else:
                matrix_of_holidays.append(0)

    # print(len(matrix_of_holidays))
    return matrix_of_holidays


def get_hours_on_holidays(sheet: Type[Worksheet], worker_cell_index) -> int:
    matrix_of_holidays = get_matrix_of_holidays(sheet)
    normalize_cells_list = get_normalize_cells_list(
        get_lines_of_working_days_dict(sheet)[worker_cell_index])

    hours_on_holidays = count_hours(list(compress(normalize_cells_list, matrix_of_holidays)))[
        'hours']

    return hours_on_holidays or 0


def get_norm_of_hours_dict(sheet: Type[Worksheet]) -> dict:
    """
    TODO
    :param sheet:
    :return:
    """
    norm_of_hours_dict = dict()

    norm_of_hours_40_cell = 'AH8'
    norm_of_hours_36_cell = 'AH9'
    norm_of_hours_28_cell = 'AH10'

    # print(norm_of_hours_40_cell.value)
    for cell in [norm_of_hours_40_cell, norm_of_hours_36_cell, norm_of_hours_28_cell]:
        if sheet[cell].value is not None:
            try:
                new_cell = float(sheet[cell].value.replace(",", "."))
            except AttributeError:
                new_cell = sheet[cell].value
            norm_of_hours_dict[cell] = new_cell

    # print(norm_of_hours_dict['AH8'])
    return norm_of_hours_dict


if __name__ == '__main__':
    # print(count_days(
    #     normalize_cells_list(
    #         get_lines_of_working_days(DEM_sheet)['B13'])))
    # print(count_hours(
    #     normalize_cells_list(
    #         get_lines_of_working_days(DEM_sheet)[15])))
    # print(write_days_to_file(DEM_sheet, 'B13'))
    print(save_sheet(DEM_sheet))
    # print(write_hours_to_file(DEM_sheet, 'B13'))
    # print(get_matrix_of_holidays(DEM_sheet))
    # print(len(get_matrix_of_holidays(DEM_sheet)),
    #       len(get_normalize_cells_list(get_lines_of_working_days(DEM_sheet)['B13'])))
    # print(get_normalize_cells_list(get_lines_of_working_days(DEM_sheet)['B13']))

    # print(get_norm_of_hours_dict(DEM_sheet))
    # print(fill_overworking_cell(DEM_sheet, 'B13'))
