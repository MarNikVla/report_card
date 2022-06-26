import re
from collections import Counter
from copy import deepcopy
from functools import cached_property

from openpyxl import load_workbook

REPORT_CARD_FILE = 'test.xlsx'
BACKUP_REPORT_CARD_FILE = f'backup_{REPORT_CARD_FILE}'

wb = load_workbook(filename=REPORT_CARD_FILE)
machinist_sheet_name, DEM_sheet_name, reason_sheet_name = wb.sheetnames

machinist_sheet, DEM_sheet, reason_of_absence_sheet = wb[machinist_sheet_name], \
                                                      wb[DEM_sheet_name], \
                                                      wb[reason_sheet_name]


class Sheet:
    def __init__(self, sheet):
        self.sheet = sheet

    @staticmethod
    def _type_of_day(cell):

        working_day = 'РД'  # Рабочий день
        weekend = 'В'  # Выходной
        short_day = 'КД'  # Короткий день
        holiday = 'П'  # Праздник

        color_standard_red = 'FFFF0000'
        color_standard_light_green = 'FF92D050'
        color_standard_orange = 'FFFFC000'

        if cell.fill.start_color.index == color_standard_red:
            return holiday
        if cell.fill.start_color.index == color_standard_light_green:
            return weekend
        if cell.fill.start_color.index == color_standard_orange:
            return short_day
        return working_day

    @cached_property
    def _work_days_matrix(self):
        first_day_of_month = self.sheet['C10']
        last_day_of_month = self.sheet['R11']
        work_days_matrix = list()
        for row in self.sheet.iter_rows(min_row=first_day_of_month.row,
                                        max_row=last_day_of_month.row,
                                        min_col=first_day_of_month.column,
                                        max_col=last_day_of_month.column):
            for cell in row:
                if cell.value not in (None, 'X'):
                    work_days_matrix.append(self._type_of_day(cell))
        # print('fsdfs', len(work_days_matrix), work_days_matrix)
        return work_days_matrix

    def __str__(self):
        return f'{self.sheet}'




class Worker(Sheet):
    def __init__(self, cell_index, sheet):
        super().__init__(sheet)
        self.cell = sheet[cell_index]

    def __str__(self):
        return f'{self.cell.value},row  {self.cell.row}, col {self.cell.column}'

    def __len__(self):
        return len(self.cells_range)

    @property
    def cells_range(self):
        cells_range = list()
        for col in self.sheet.iter_rows(min_row=self.cell.row,
                                        max_row=self.cell.row + 1,
                                        min_col=self.cell.column + 1,
                                        max_col=self.cell.column + 16):
            for cell in col:
                if cell.value not in (None, 'Х'):
                    cells_range.append(cell.value)
        return cells_range

    @cached_property
    def is_28_hours_week(self):
        color_standard_orange = 'FFFFC000'
        return self.cell.fill.start_color.index == color_standard_orange

    @property
    def _work_days_matrix(self):
        return super(Worker, self)._work_days_matrix[:len(self.cells_range) + 1]

    @cached_property
    def _normalize_workdays(self):
        days_to_remove = ['ОТ', 'У', 'ДО', 'Б', 'К', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ', 'НОД']
        normalize_workdays = deepcopy(self._work_days_matrix)
        for index, cell in enumerate(self.cells_range):
            if cell in days_to_remove:
                normalize_workdays[index] = 0
        return normalize_workdays

    @cached_property
    def counter_of_days(self):
        return Counter(self._normalize_workdays)

    @cached_property
    def norm_of_hours(self):
        counter = self.counter_of_days
        duration_of_day = 8
        if self.is_28_hours_week:
            duration_of_day = 5.6
        duration_of_short_day = duration_of_day - 1
        norm_of_hours = counter['РД'] * duration_of_day + counter['КД'] * duration_of_short_day
        return norm_of_hours

    def get_weekends(self):
        return self.counter_of_days['В']

    def get_vacation_days(self):
        return self.counter_of_days['OT']

    def get_medical_days(self):
        return self.counter_of_days['Б']

    def get_NOD_days(self):
        return self.counter_of_days['НОД']

    def get_other_days_off(self):
        return sum(
            [self.counter_of_days[key] for key in ['ОВ', 'У', 'ДО', 'К', 'ПР', 'Р', 'ОЖ', 'ОЗ', 'Г', 'НН', 'НБ']])

    # def get_attendance_days(self):
    #     attendance_days = sum(
    #         counter.values()) - (absence_days + vacation_days + medical_days +
    #                              other_absence_days + absence_paid_days)
    #     return self.counter_of_days['В']


    @property
    def _prepared_range(self):
        new_cells_list = list()
        for cell in self.cells_range:
            if cell is not None:
                try:
                    new_cell = float(cell.replace(",", "."))
                except ValueError:
                    # remove whitespaces
                    new_cell = re.sub(r'\s+', '', cell)
                new_cells_list.append(new_cell)
        return new_cells_list

    @staticmethod
    def count_hours(cells_range):
        day_hours = 0
        night_hours = 0
        counter_of_hours = Counter(cells_range)
        for i in counter_of_hours.keys():
            if isinstance(i, float):
                day_hours += i * counter_of_hours[i]
            elif i == '8/20':
                day_hours += 12 * counter_of_hours[i]
            elif i in ['20/', '20/24']:
                day_hours += 4 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i in ['/820/24', '0/820/', '/820/']:
                day_hours += 12 * counter_of_hours[i]
                night_hours += 8 * counter_of_hours[i]
            elif i == '820/':
                day_hours += 12 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i == '420/':
                day_hours += 8 * counter_of_hours[i]
                night_hours += 2 * counter_of_hours[i]
            elif i in ['0/8', '/8']:
                day_hours += 8 * counter_of_hours[i]
                night_hours += 6 * counter_of_hours[i]
        # d = {'night_hours': night_hours, 'all_hours': all_hours}
        return {'night_hours': night_hours, 'day_hours': day_hours}

    def get_day_hours(self):
        count_day_hours = self.count_hours(self._prepared_range)['day_hours']
        return count_day_hours

    def get_night_hours(self):
        count_night_hours = self.count_hours(self._prepared_range)['night_hours']
        return count_night_hours

    def get_holidays(self):
        holidays_range = list()
        for i, cell in enumerate(self._prepared_range):
            if worker._normalize_workdays[i] == 'П':
                holidays_range.append(cell)
        count_holiday_hours = self.count_hours(holidays_range)['day_hours']
        return count_holiday_hours

    def get_overwork(self):
        return self.get_day_hours() - self.norm_of_hours

    # def fill_sheet(self):
    #     offset_row = 0
    #     offset_column_attendance = 17
    #     offset_column_absence = 22
    #     offset_column_vacation = 23
    #     offset_column_medical = 24
    #     offset_column_other_absence = 25
    #     offset_column_other_absence_paid = 26
    #     self.sheet[self.cell].offset(row=offset_row, column=offset_column_attendance).value =
    #     pass

worker = Worker('B13', DEM_sheet)
worker_women = Worker('B35', DEM_sheet)

print(worker)
# print(len(worker.cells_range),worker.cells_range)
# print(len(worker.work_days_matrix), worker.work_days_matrix)
# print(len(worker.normalize_workdays), worker.normalize_workdays)
print(worker.norm_of_hours)
print(worker.get_day_hours())
print(worker.get_night_hours())
print(worker.get_overwork())

